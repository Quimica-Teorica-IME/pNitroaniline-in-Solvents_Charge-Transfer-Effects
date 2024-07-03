import os
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Alignment

def execute_command(command, directory, subdir):
    print(f"Executando comando '{command}' em '{subdir}'")
    process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True, cwd=directory)
    stdout, stderr = process.communicate()
    if process.returncode != 0:
        try:
            error_message = stderr.decode('latin-1')
        except UnicodeDecodeError:
            error_message = stderr.decode('cp1252')
        print(f"Erro ao executar '{command}' em {subdir}: {error_message}")
        return False
    return True

def execute_theodore_analyze_tden(directory, subdir):
    return execute_command('theodore analyze_tden', directory, subdir)

def execute_theodore_plot_om_bars(directory, subdir):
    print(f"Executando comando 'theodore plot_om_bars' em '{subdir}'")
    command = 'theodore plot_om_bars'
    process = subprocess.Popen(command, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True, cwd=directory)
    input_text = '''OmFrag.txt
tden_summ.txt
7.000000
ACT
cyan
1 3
2 3

NCT
orange
2 1
3 1

RCT
gray
1 2
3 2

ALE
blue
3 3

NLE
red
1 1

RLE
yellow
2 2





'''
    stdout, stderr = process.communicate(input=input_text.encode('utf-8'))
    if process.returncode != 0:
        try:
            error_message = stderr.decode('latin-1')
        except UnicodeDecodeError:
            error_message = stderr.decode('cp1252')
        print(f"Erro ao executar 'theodore plot_om_bars' em {subdir}: {error_message}")
        return False
    return True

def execute_pdflatex(directory, subdir):
    return execute_command('pdflatex Om_bars.tex', directory, subdir)

def extract_om_bar_data(directory):
    om_bar_path = os.path.join(directory, 'Om_bar_data.txt')
    data = []

    if not os.path.exists(om_bar_path):
        print(f"Arquivo {om_bar_path} não encontrado.")
        return data

    with open(om_bar_path, 'r') as file:
        lines = file.readlines()
        
        for line in lines[1:]:
            columns = line.strip().split()
            state = columns[1]
            energy = columns[2]
            f_value = columns[3]
            act = columns[4]
            nct = columns[5]
            rct = columns[6]
            ale = columns[7]
            nle = columns[8]
            rle = columns[9]
            data.append((state, energy, f_value, act, nct, rct, ale, nle, rle))

    return data

def create_excel_file(base_directory, all_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "OM Bars Data"
    col = 1

    for directory_name, data in all_data.items():
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+8)
        ws.cell(row=1, column=col).value = directory_name
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

        headers = ['State', 'Energy', 'f', 'ACT', 'NCT', 'RCT', 'ALE', 'NLE', 'RLE']
        for i, header in enumerate(headers):
            ws.cell(row=2, column=col + i, value=header)

        for row_idx, row_data in enumerate(data, start=3):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=row_idx, column=col + col_idx, value=value)

        col += 10

    excel_path = os.path.join(base_directory, 'OM_Bars_Data.xlsx')
    wb.save(excel_path)
    print(f"Arquivo Excel salvo em {excel_path}")

def process_directories_in_current_directory():
    try:
        current_directory = os.path.abspath(os.path.dirname(__file__))
    except NameError:
        current_directory = os.getcwd()

    subdirectories = [subdir for subdir in os.listdir(current_directory) if os.path.isdir(os.path.join(current_directory, subdir))]
    subdirectories.sort()
    all_data = {}

    for subdir in subdirectories:
        subdir_path = os.path.join(current_directory, subdir)
        print(f"Entrando no diretório: {subdir}")

        if execute_theodore_analyze_tden(subdir_path, subdir):
            if execute_theodore_plot_om_bars(subdir_path, subdir):
                if execute_pdflatex(subdir_path, subdir):
                    data = extract_om_bar_data(subdir_path)
                    all_data[subdir] = data

        print("--------------------------------------------------\n")

    create_excel_file(current_directory, all_data)

if __name__ == "__main__":
    process_directories_in_current_directory()
